from odoo import models, fields, api, _
from odoo.exceptions import UserError

import logging
import csv
import os
from datetime import datetime
import re
import io
import base64
import xlsxwriter
from io import BytesIO
from odoo.exceptions import ValidationError
from odoo.tools.float_utils import float_compare

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

_logger = logging.getLogger(__name__)


class ProductTemplate(models.Model):
    _inherit = 'product.template'
    _last_imported_template_id = None

    export_uom = fields.Char('Export UOM')
    export_po_uom = fields.Char('Export Po UOM')
    export_uom_category = fields.Char('Export UOM Category')
    export_barcode_uom_ids = fields.Char('Export barcode uom ids')
    export_barcode_line_uom = fields.Char('Export barcode line uom')
    export_barcode_sale_price = fields.Char('Export barcode sale price')
    export_barcode_description = fields.Char('Export barcode description')
    export_vendor_list_po_uom = fields.Char('Export Purchase unit of measure')
    processed_uom = fields.Boolean(default=False)

    def _get_factor_inv_from_uom_name(self, uom_name):
        if not uom_name:
            return 1.0

        match = re.search(r'\((\d+)\s*x\)', uom_name, re.IGNORECASE)
        if match:
            return float(match.group(1))

        return 1.0

    @api.model
    def create(self, vals):
        barcode_to_validate = []

        main_barcode = vals.get("barcode")
        if main_barcode:
            barcode_to_validate.append(main_barcode)

        if main_barcode:
            existing_product = self.env["product.template"].search([
                ("barcode", "=", main_barcode)
            ], limit=1)
            if existing_product:
                raise ValidationError(
                    ("Barcode '%s' already exists for product '%s', Duplicate Barcode ") %
                    (main_barcode, existing_product.display_name)
                )
                
        uom_barcode_raw = vals.get("export_barcode_uom_ids")
        uom_barcodes = []
        if uom_barcode_raw:
            if isinstance(uom_barcode_raw, str):
                uom_barcodes = [x.strip() for x in uom_barcode_raw.split(",") if x.strip()]
        barcode_to_validate.extend(uom_barcodes)


        for bc in barcode_to_validate:
            existing_uom = self.env["product.barcode.uom"].search([
                ("barcode", "=", bc)
            ], limit=1)
            if existing_uom:
                raise ValidationError(
                    ("Barcode '%s' already exists for product '%s' (UOM: %s), Duplicate Barcode ") %
                    (bc, existing_uom.product_id.display_name, existing_uom.uom_id.name)
                )

        product_name = vals.get('name')
        existing_product = False

        if product_name:
            # Normal product row → search by name
            existing_product = self.env['product.template'].search([
                ('name', 'ilike', product_name)
            ], limit=1)

        else:
            last_id = self._last_imported_template_id
            if last_id:
                existing_product = self.env['product.template'].browse(last_id)

        category_name = vals.get('export_uom_category')
        uom_category = None

        if category_name:
            uom_category = self.env['uom.category'].search([('name', 'ilike', category_name)], limit=1)
            if not uom_category:
                uom_category = self.env['uom.category'].create({'name': category_name})
                _logger.info(f"Created new UOM Category: {category_name}")
            vals['uom_category_id'] = uom_category.id

            # Create export_uom if not exists
            # uom_name = vals.get('export_uom')
            # if uom_name:
            #     uom = self.env['uom.uom'].search([
            #         ('category_id', '=', uom_category.id),
            #         ('name', 'ilike', uom_name)
            #     ], limit=1)
            #     if not uom:
            #         uom = self.env['uom.uom'].create({
            #             'name': uom_name,
            #             'category_id': uom_category.id,
            #             'uom_type': 'reference',
            #             'factor': 1.0,
            #         })
            #     vals['uom_id'] = uom.id

            uom_name = vals.get('export_uom')
            if uom_name:
                uom = self.env['uom.uom'].search([
                    ('category_id', '=', uom_category.id),
                    ('name', 'ilike', uom_name)
                ], limit=1)
                if not uom:
                    # Check if reference already exists in this category
                    ref_uom = self.env['uom.uom'].search([
                        ('category_id', '=', uom_category.id),
                        ('uom_type', '=', 'reference')
                    ], limit=1)

                    uom_type = 'reference' if not ref_uom else 'bigger'

                    uom = self.env['uom.uom'].create({
                        'name': uom_name,
                        'category_id': uom_category.id,
                        'uom_type': uom_type,
                        'factor': 1.0,
                        'factor_inv': 1.0 if uom_type == 'reference' else 2.0,  # adjust as needed
                    })
                vals['uom_id'] = uom.id

            # Create export_po_uom if not exists
            po_uom_name = vals.get('export_po_uom')
            if po_uom_name:
                po_uom = self.env['uom.uom'].search([
                    ('category_id', '=', uom_category.id),
                    ('name', 'ilike', po_uom_name)
                ], limit=1)
                if not po_uom:
                    po_uom = self.env['uom.uom'].create({
                        'name': po_uom_name,
                        'category_id': uom_category.id,
                        'uom_type': 'bigger',
                        'factor_inv': 2.0,
                    })
                vals['uom_po_id'] = po_uom.id

        # === Create Product ===
        if existing_product:
            product = existing_product
            product.ctn_qty = category_name
            _logger.info(f"Product '{product_name}' exists (ID: {product.id}), updating vendor lines and barcodes.")
            if vals.get("seller_ids"):
                for _, _, seller_vals in vals["seller_ids"]:
                    uom_name = vals.get("export_vendor_list_po_uom")
                    if not uom_name:
                        continue

                    uom = self.env["uom.uom"].search([
                        ("category_id", "=", product.uom_category_id.id),
                        ("name", "ilike", uom_name)
                    ], limit=1)

                    if not uom:
                        uom = self.env["uom.uom"].create({
                            "name": uom_name,
                            "category_id": product.uom_category_id.id,
                            "uom_type": "bigger",
                            "factor_inv": 1.0,
                        })

                    seller_vals["product_uom"] = uom.id
                    product.seller_ids = [(0, 0, seller_vals)]
        else:
            product = super().create(vals)
            product.ctn_qty = category_name
            if not self.env.context.get("child_barcode_product"):
                if product_name and not vals.get('export_barcode_uom_ids'):
                    self.__class__._last_imported_template_id = product.id
            _logger.info(f"Created new product '{product_name}' with ID {product.id}")

        main_barcode = vals.get("barcode")

        if (main_barcode and not self.env.context.get("child_barcode_product")   
            and not existing_product and not vals.get('export_barcode_uom_ids')):
            exists = self.env['product.barcode.uom'].search([
                ('product_id', '=', product.id),
                ('barcode', '=', main_barcode),
            ], limit=1)

            if not exists:
                res = self.env['product.barcode.uom'].create({
                    'product_id': product.id,
                    'barcode': main_barcode,
                    'uom_id': product.uom_id.id,
                    'uom_category_id': product.uom_category_id.id,
                    'sale_price': product.list_price or 0.0,
                    'description': product.name or "",
                })
                _logger.info(f"Created barcode_uom record for parent: {main_barcode}")
            
        # === Handle Barcodes + UOMs + Sale Prices ===
        barcode_string = vals.get('export_barcode_uom_ids')
        barcode_uom_names = vals.get('export_barcode_line_uom')
        barcode_sale_price = vals.get('export_barcode_sale_price')
        barcode_description = vals.get('export_barcode_description')

        if barcode_string and barcode_uom_names:
            try:
                barcode_list = [b.strip() for b in barcode_string.split(',') if b.strip()]
                uom_name_list = [u.strip() for u in barcode_uom_names.split(',') if u.strip()]
                description_list = [u.strip() for u in barcode_description.split(',') if u.strip()]
                price_list = [float(p.strip()) if p.strip() else 0.0 for p in (barcode_sale_price or '').split(',')]

                for idx, barcode in enumerate(barcode_list):
                    uom_name = uom_name_list[idx] if idx < len(uom_name_list) else None
                    description = description_list[idx] if idx < len(description_list) else None
                    price = price_list[idx] if idx < len(price_list) else 0.0
                    uom_id = False

                    if uom_name:
                        existing_uom = self.env['uom.uom'].search([
                            ('category_id', '=', product.uom_category_id.id),
                            ('name', 'ilike', uom_name)
                        ], limit=1)

                        if not existing_uom:
                            # Check if any reference UOM exists in this category
                            ref_uom = self.env['uom.uom'].search([
                                ('category_id', '=', product.uom_category_id.id),
                                ('uom_type', '=', 'reference'),
                            ], limit=1)

                            uom_type = 'reference' if not ref_uom else 'bigger'
                            factor_inv = self._get_factor_inv_from_uom_name(uom_name)

                            existing_uom = self.env['uom.uom'].create({
                                'name': uom_name,
                                'category_id': product.uom_category_id.id,
                                'uom_type': uom_type,
                                # 'factor': 1.0,
                                'factor': 1.0 / factor_inv,
                                'factor_inv': factor_inv,
                            })
                            _logger.info(f"Created new UOM '{existing_uom}','{uom_name}' with type '{uom_type}' in category ID {product.uom_category_id.id}")

                        uom_id = existing_uom.id

                    # Check if a product exists for this barcode
                    existing_product = self.env['product.product'].search([('barcode', '=', barcode)], limit=1)
                    if not existing_product:
                        # Create new product.product for this barcode
                        new_template_vals = {
                            'name': description or barcode,
                            'barcode': barcode,
                            'available_in_pos': True,
                            'active': False,
                            'list_price': price,
                            'uom_id': uom_id or product.uom_id.id,
                            'uom_po_id': product.uom_po_id.id,
                            'uom_category_id': product.uom_category_id.id,
                        }
                        new_template = self.with_context(child_barcode_product=True).env['product.template'].create(new_template_vals)
                        _logger.info(f"Created new product '{new_template.name}' for barcode {barcode}")
                        existing_product = new_template.product_variant_id
                        
                    if uom_id:
                        exists = self.env['product.barcode.uom'].search([
                            ('product_id', '=', product.id),
                            ('uom_id', '=', uom_id),
                            ('barcode', '=', barcode),
                        ], limit=1)

                        if not exists:
                            res = self.env['product.barcode.uom'].create({
                                'product_id': product.id,
                                'description': description,
                                'barcode': barcode,
                                'uom_id': uom_id,
                                'uom_category_id': product.uom_category_id.id,
                                'sale_price': price,
                            })
                        else:
                            _logger.info(f"Duplicate barcode '{barcode}' found for product {product.id}; skipping.")

            # except Exception as e:
            #     _logger.error(f"Error while processing barcodes: {e}")
            except psycopg2.IntegrityError:
                self.env.cr.rollback()
                raise ValidationError(
                    _("The barcode '%s' is already used by another product.\n"
                      "Barcodes must be unique. Please correct your Excel file.")
                    % barcode
                )

            except Exception as e:
                self.env.cr.rollback()
                raise ValidationError(
                    _("Barcode processing failed:\n%s") % str(e)
                )

        return product

    def action_set_product(self):
        for rec in self:
            count_invalid_children0 = self.env['product.template'].search_count([
                ('parent_template_id', '!=', rec.id),
                ('active', '=', True),
                ('unit_qty', '<', 1),
            ])
        _logger.info(f"[{rec.id}] Found {count_invalid_children0} product.template record(s) with unit_qty < 1 and active=True (excluding self).")
        Param = self.env['ir.config_parameter']
        batch_size = 10000

        # Get last offset, default 0
        last_offset = int(Param.sudo().get_param('uom_update_last_offset', default=0))

        # Fetch next batch of products
        products = self.env['product.template'].search([], offset=last_offset, limit=batch_size)
        if not products:
            # If no more records, reset offset and start again
            last_offset = 0
            products = self.env['product.template'].search([], offset=last_offset, limit=batch_size)

        product_ids = tuple(products.ids)
        if not product_ids:
            _logger.info("No product templates found for update.")
            return

        start_num = last_offset + 1
        end_num = last_offset + len(product_ids)

        # Print which range is processed
        _logger.info("Processing product templates from %s to %s", start_num, end_num)

        UoMModel = self.env['uom.uom']
        UoMCategoryModel = self.env['uom.category']

        updated_less_than_1 = 0
        updated_equal_to_1 = 0
        updated_bigger = 0
        updated_less_than_10 = 0
        updated_equal_to_10 = 0

        for rec in products:
            unit_qty = rec.unit_qty
            carton_qty = rec.ctn_qty
            if not carton_qty:
                continue  # Skip if qty not set

            qty_str = str(int(carton_qty)) if carton_qty == int(carton_qty) else str(carton_qty)

            # ---------- PARENT TEMPLATE ----------
            if rec.id == rec.parent_template_id.id:
                _logger.info(f"\n\n[PARENT PRODUCT] Processing template {rec.id} with carton qty: {qty_str}")

                uom_category_name = qty_str
                uom_name = "UNIT"

                if rec.barcode and rec.barcode.startswith('99'):
                    suffix = rec.barcode[2:]
                    if rec.default_code and suffix.startswith(rec.default_code):
                        uom_category_name = "WEIGHT"
                        uom_name = "KG"

                # Find or create UoM Category
                uom_category = UoMCategoryModel.search([('name', '=', uom_category_name)], limit=1)
                if not uom_category:
                    uom_category = UoMCategoryModel.create({
                        'name': uom_category_name,
                        'is_pos_groupable': True,
                    })
                    _logger.info(f"Created UoM category: {uom_category.name}")

                # Find or create reference UoM in that category
                reference_uom = UoMModel.search([
                    ('category_id', '=', uom_category.id),
                    ('name', 'ilike', uom_name),
                ], limit=1)
                uom_type = 'reference' if not reference_uom else 'bigger'
                if not reference_uom:
                    reference_uom = UoMModel.create({
                        'name': uom_name,
                        'category_id': uom_category.id,
                        'uom_type': uom_type,
                        'rounding': 0.01,
                        'factor': 1.0,
                    })
                    _logger.info(f"Created reference UoM: {reference_uom.name}")

                rec.write({
                    'uom_category_id': uom_category.id,
                })

                # Step 2: Update uom_id and uom_po_id using direct SQL (manual way)
                self.env.cr.execute("""
                    UPDATE product_template
                    SET 
                        uom_id = %s,
                        uom_po_id = %s
                    WHERE id = %s
                """, (
                    reference_uom.id,
                    reference_uom.id,
                    rec.id
                ))

                _logger.info(f"[PARENT] Updated template ID {rec.id} with reference UoM")

            # ---------- CHILD TEMPLATE ----------
            else:
                _logger.info(f"\n\n[CHILD PRODUCT] Processing child template {rec.id}")

                if unit_qty < 1:
                    _logger.info(f"[CHILD] unit_qty < 1 for template {rec.id}, skipping parent linking...")

                    # Set as self-parent
                    rec.parent_template_id = rec.id
                    rec.unit_qty = 1
                    rec.ctn_qty = 1

                    # Search or create '1' category
                    unit_category = UoMCategoryModel.search([('name', '=', '1')], limit=1)
                    if not unit_category:
                        unit_category = UoMCategoryModel.create({'name': '1'})

                    # Search or create 'Unit' UoM in that category
                    unit_uom = UoMModel.search([
                        ('category_id', '=', unit_category.id),
                        ('factor', '=', 1.0),
                        ('uom_type', '=', 'reference'),
                    ], limit=1)

                    if not unit_uom:
                        unit_uom = UoMModel.create({
                            'name': 'Unit',
                            'category_id': unit_category.id,
                            'uom_type': 'reference',
                            'rounding': 0.01,
                            'factor': 1.0,
                        })
                        _logger.info(f"[CHILD-SKIP] Created new 'Unit' UoM in category ID {unit_category.id}")

                    # Update the product_template's UoM and category
                    if unit_uom:
                        self.env.cr.execute("""
                            UPDATE product_template
                            SET 
                                uom_id = %s,
                                uom_po_id = %s,
                                uom_category_id = %s
                            WHERE id = %s
                        """, (unit_uom.id, unit_uom.id, unit_category.id, rec.id))
                        _logger.info(f"[CHILD-SKIP] Set as independent with 'Unit' UoM")

                        rows_child_skip = self.env.cr.rowcount
                        _logger.info(f"[CHILD-SKIP] Set as independent with 'Unit' UoM (Updated {rows_child_skip} record(s))")
                        updated_less_than_10 += rows_child_skip
                        updated_less_than_1 += 1
                    continue  # Skip rest of child logic

                # Normal flow for child with unit_qty >= 1
                parent = self.env['product.template'].browse(rec.parent_template_id.id)
                if not parent.uom_category_id:
                    _logger.warning(f"[CHILD] Skipping child {rec.id}: Parent has no UoM category")
                    continue

                uom_category_id = parent.uom_category_id
                factor = 1.0 / unit_qty if unit_qty else 0.0

                # Special case: if unit_qty == 1, use Unit in parent's category
                if unit_qty == 1:
                    unit_uom = UoMModel.search([
                        ('category_id', '=', uom_category_id.id),
                        ('factor', '=', 1.0),
                        ('uom_type', '=', 'reference'),
                    ], limit=1)

                    if not unit_uom:
                        unit_uom = UoMModel.create({
                            'name': 'Unit',
                            'category_id': uom_category_id.id,
                            'uom_type': 'reference',
                            'rounding': 0.01,
                            'factor': 1.0,
                        })
                        _logger.info(f"[CHILD] Created reference Unit UoM in parent's category {uom_category_id.name}")

                    self.env.cr.execute("""
                        UPDATE product_template
                        SET uom_id = %s,
                            uom_category_id = %s,
                            uom_po_id = %s
                        WHERE id = %s
                    """, (unit_uom.id, uom_category_id.id, unit_uom.id, rec.id))
                    _logger.info(f"[CHILD] Updated template ID {rec.id} with reference Unit UoM from parent's category")
                    
                    rows_equal_to_1 = self.env.cr.rowcount
                    _logger.info(f"[CHILD] Updated template ID {rec.id} with reference Unit UoM from parent's category (Updated {rows_equal_to_1} record(s))")
                    updated_equal_to_10 += rows_equal_to_1
                    updated_equal_to_1 += 1

                    continue  # Skip rest of child logic since handled

                # Check if child-specific UoM already exists
                existing_bigger_uom = UoMModel.search([
                    ('category_id', '=', uom_category_id.id),
                    ('uom_type', '=', 'bigger'),
                    ('factor', '=', factor)
                ], limit=1)
                if not existing_bigger_uom:
                    if unit_qty == parent.ctn_qty:
                        new_uom_name = f"CTN ({int(unit_qty)}X)"
                    elif unit_qty < parent.ctn_qty and unit_qty != 1:
                        new_uom_name = f"OUTER ({int(unit_qty)}X)"
                    else:
                        new_uom_name = f"{int(unit_qty)}X"

                    new_uom = UoMModel.create({
                        'name': new_uom_name,
                        'category_id': uom_category_id.id,
                        'uom_type': 'bigger',
                        'rounding': 0.01,
                        'factor_inv': unit_qty,
                    })
                    _logger.info(f"[CHILD] Created bigger UoM: {new_uom.name}")
                else:
                    new_uom = existing_bigger_uom
                    _logger.info(f"[CHILD] Found existing bigger UoM: {new_uom.name}")

                self.env.cr.execute("""
                    UPDATE product_template
                    SET uom_id = %s,
                        uom_category_id = %s,
                        uom_po_id = %s
                    WHERE id = %s
                """, (new_uom.id, uom_category_id.id, new_uom.id, rec.id))
                _logger.info(f"[CHILD] Updated template ID {rec.id} with bigger UoM")
                rows_bigger = self.env.cr.rowcount
                _logger.info(f"[CHILD] Updated template ID {rec.id} with bigger UoM (Updated {rows_bigger} record(s))")
                updated_bigger += rows_bigger


        new_offset = last_offset + batch_size
        total_products = self.env['product.template'].search_count([])
        if new_offset >= total_products:
            new_offset = 0  # reset if end reached
        Param.sudo().set_param('uom_update_last_offset', new_offset)

        for rec in self:
            count_invalid_children1 = self.env['product.template'].search_count([
                ('parent_template_id', '!=', rec.id),
                ('active', '=', True),
                ('unit_qty', '<', 1),
            ])

    def action_generate_barcode_uom_lines(self):
        Param = self.env['ir.config_parameter']
        BarcodeUom = self.env['product.barcode.uom']
        ProductSupplier = self.env['product.supplierinfo']
        batch_size = 7000
        failed_lines = []

        products = self.env['product.template'].search([('processed_uom', '=', False)], limit=batch_size)

        if not products:
            _logger.info("All products already processed. Nothing to update.")
            return

        for template in products:
            is_child = template.id != template.parent_template_id.id

            if is_child and template.parent_template_id and template.parent_template_id.active:
                parent = template.parent_template_id
                comp_id = template.company_id.id or self.env.user.company_id.id
                chatter_lines = []

                # Get child product quants
                self.env.cr.execute("""
                    SELECT location_id, quantity
                    FROM stock_quant
                    WHERE product_id = %s
                """, (template.product_variant_id.id,))
                child_quants = self.env.cr.fetchall()

                for loc_id, qty in child_quants:
                    adjusted_qty = template.unit_qty * qty if template.unit_qty else qty

                    # Parent old qty
                    self.env.cr.execute("""
                        SELECT quantity
                        FROM stock_quant
                        WHERE product_id = %s
                          AND location_id = %s
                          AND company_id = %s
                    """, (parent.product_variant_id.id, loc_id, comp_id))
                    res = self.env.cr.fetchone()
                    parent_old_qty = res[0] if res else 0.0

                    # Update or insert
                    self.env.cr.execute("""
                        UPDATE stock_quant
                        SET quantity = quantity + %s
                        WHERE product_id = %s
                          AND location_id = %s
                          AND company_id = %s
                    """, (adjusted_qty, parent.product_variant_id.id, loc_id, comp_id))

                    if self.env.cr.rowcount == 0:
                        self.env.cr.execute("""
                            INSERT INTO stock_quant (product_id, location_id, quantity, reserved_quantity, company_id)
                            VALUES (%s, %s, %s, 0, %s)
                        """, (parent.product_variant_id.id, loc_id, adjusted_qty, comp_id))
                        parent_old_qty = 0.0

                    # Sync supplier UoM
                    for seller in template.seller_ids:
                        if template.uom_po_id.id != seller.product_uom.id:
                            # seller.product_uom = template.uom_po_id
                            seller.write({'product_uom': template.uom_po_id.id})

                    # Chatter logging
                    location_rec = self.env['stock.location'].browse(loc_id)
                    if location_rec.usage == 'internal':
                        chatter_lines.append(
                            f"Location: <b>{location_rec.display_name}</b><br/>"
                            f"Parent Qty (Before): {parent_old_qty}<br/>"
                            f"Child Qty × Unit Qty: {adjusted_qty} "
                            f"({qty} × {template.unit_qty if template.unit_qty else 1})<br/>"
                            f"<b>Total Qty (After):</b> {parent_old_qty + adjusted_qty}"
                        )

                # Post chatter
                if chatter_lines:
                    parent.message_post(
                        body="<br/><br/>".join(chatter_lines),
                        subtype_xmlid="mail.mt_note"
                    )

                # Ensure child seller UoM is correct
                for seller in template.seller_ids:
                    if template.uom_po_id.id != seller.product_uom.id:
                        # seller.product_uom = template.uom_po_id
                        seller.write({'product_uom': template.uom_po_id.id})
                template.write({'processed_uom': True})
                continue

            # -----------------------
            # Parent Template Section
            # -----------------------
            # _logger.info(f"\n\n[PARENT] Processing template: {template.name} (ID: {template.id})")

            # Update parent seller UoM
            for seller in template.seller_ids:
                if seller.product_uom.id != template.uom_po_id.id:
                    # seller.product_uom = template.uom_po_id
                    seller.write({'product_uom': template.uom_po_id.id})

            # Add barcode for parent if missing
            if template.barcode:
                if not BarcodeUom.search([('barcode', '=', template.barcode)], limit=1):
                    try:
                        template.barcode_uom_ids = [(0, 0, {
                            'product_id': template.id,
                            'uom_category_id': template.uom_category_id.id,
                            'description': template.name,
                            'uom_id': template.uom_id.id,
                            'barcode': template.barcode,
                            'sale_price': template.list_price,
                        })]
                        if template.multi_barcode_ids:
                            for barcode in template.multi_barcode_ids:
                                template.barcode_uom_ids = [(0, 0, {
                                    'product_id': template.id,
                                    'uom_category_id': template.uom_category_id.id,
                                    'description': template.name,
                                    'uom_id': template.uom_id.id,
                                    'barcode': barcode.multi_barcode,
                                    'sale_price': template.list_price,
                                })]
                        _logger.info(f"[PARENT] Added barcode line for parent: {template.barcode}")
                    except ValidationError as e:
                        template.write({'processed_uom': True})
                        failed_lines.append({
                            'product_name': template.name,
                            'barcode': template.barcode,
                            'uom': template.uom_id.name if template.uom_id else '',
                            'error': str(e),
                        })

            # Process child templates
            child_templates = self.env['product.template'].search([
                ('parent_template_id', '=', template.id),
                ('id', '!=', template.id)
            ])

            for child in child_templates:
                if not child.barcode:
                    template.write({'processed_uom': True})
                    continue

                # Check if child barcode already exists in parent's lines
                if not template.barcode_uom_ids.filtered(lambda b: b.barcode == child.barcode):
                    try:
                        barcode_lines = [(0, 0, {
                            'product_id': template.id,  # if you want barcodes on parent
                            'uom_category_id': child.uom_category_id.id,
                            'description': child.name,
                            # 'description': f"{template.name} X {child.uom_id.name}",
                            'uom_id': child.uom_id.id,
                            'barcode': child.barcode,
                            'sale_price': child.list_price,
                        })]

                        for barcode in child.multi_barcode_ids:
                            barcode_lines.append((0, 0, {
                                'product_id': template.id,
                                'uom_category_id': child.uom_category_id.id,
                                'description': child.name,
                                # 'description': f"{template.name} X {child.uom_id.name}",
                                'uom_id': child.uom_id.id,
                                'barcode': barcode.multi_barcode,
                                'sale_price': child.list_price,
                            }))

                        template.barcode_uom_ids = barcode_lines
                        _logger.info(f"[CHILD] Added barcode line for child: {child.barcode}")
                    except ValidationError as e:
                        template.write({'processed_uom': True})
                        failed_lines.append({
                            'product_name': child.name,
                            'barcode': child.barcode,
                            'uom': child.uom_id.name if child.uom_id else '',
                            'error': str(e),
                        })


                # Sync child seller UoM and copy to parent if missing
                for seller in child.seller_ids:
                    if seller.product_uom.id != child.uom_po_id.id:
                        # seller.product_uom = child.uom_po_id
                        seller.write({'product_uom': child.uom_po_id.id})

                    matched = template.seller_ids.filtered(
                        lambda s: s.name.id == seller.name.id and
                                  s.min_qty == seller.min_qty and
                                  float_compare(s.price, seller.price, precision_digits=2) == 0 and
                                  s.delay == seller.delay and
                                  s.product_code == seller.product_code and
                                  s.product_name == seller.product_name
                    )

                    if matched:
                        for match in matched:
                            if match.product_uom.id != child.uom_po_id.id:
                                # match.product_uom = child.uom_po_id
                                match.write({'product_uom': child.uom_po_id.id})
                                _logger.info(
                                    f"[SELLER-UOM] Updated UOM for seller {match.name.display_name} on parent template"
                                )
                    elif template.purchase_ok:
                        try:
                            ProductSupplier.create({
                                'name': seller.name.id,
                                'product_tmpl_id': template.id,
                                'min_qty': seller.min_qty,
                                'price': seller.price,
                                'delay': seller.delay,
                                'product_code': seller.product_code,
                                'product_name': seller.product_name,
                                'product_uom': seller.product_uom.id,
                            })
                            _logger.info(f"[SELLER] Copied seller {seller.name.display_name} from child to parent")
                        except Exception as e:
                            template.write({'processed_uom': True})
                            failed_lines.append({
                                'product_name': child.name,
                                'seller': seller.name.name,
                                'error': str(e),
                            })
            template.write({'processed_uom': True})

        _logger.info(f"[LINES] Created : {failed_lines}")

        if failed_lines:
            return self._generate_failed_report(template, failed_lines)

    def _generate_failed_report(self, template, failed_lines):
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        sheet = workbook.add_worksheet("Failed Barcode Lines")

        headers = ['Product Name', 'Barcode', 'UOM', 'Error Message']
        for col, header in enumerate(headers):
            sheet.write(0, col, header)

        for row, data in enumerate(failed_lines, start=1):
            sheet.write(row, 0, data['product_name'])
            sheet.write(row, 1, data['barcode'])
            sheet.write(row, 4, data['uom'])
            sheet.write(row, 5, data['error'])

        workbook.close()
        output.seek(0)

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': 'failed_barcode_lines.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': template._name,
            'res_id': template.id,
        })
        _logger.info(f"[ATTACHMENT] Created: {attachment.name} (ID: {attachment.id})")
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_update_product_history(self):
        try:
            Param = self.env['ir.config_parameter'].sudo()
            batch_size = 10000

            # Get last offset, default 0
            last_offset = int(Param.get_param('product_update_last_offset') or 0)

            products = self.env['product.template'].search([], offset=last_offset, limit=batch_size)
            if not products:
                last_offset = 0
                products = self.env['product.template'].search([], offset=last_offset, limit=batch_size)

            if not products:
                _logger.info("No product templates found for update.")
                return

            product_ids = tuple(products.ids)
            start_num = last_offset + 1
            end_num = last_offset + len(product_ids)
            _logger.info("Processing product templates from %s to %s", last_offset + 1, last_offset + len(products))

            # ---------------- SALE ORDER LINE ----------------
            # self.env.cr.execute("""
            #     UPDATE sale_order_line sol
            #     SET product_id = pp_parent.id
            #     FROM product_product pp
            #     JOIN product_template pt ON pp.product_tmpl_id = pt.id
            #     JOIN product_template pt_parent ON pt.parent_template_id = pt_parent.id
            #     JOIN product_product pp_parent ON pp_parent.product_tmpl_id = pt_parent.id
            #     WHERE sol.product_id = pp.id
            #       AND pp_parent.id IS NOT NULL
            #       AND sol.product_id != pp_parent.id
            #       AND pt.id IN %s
            #       AND pt.uom_category_id IS NULL
            #       AND pt.active = TRUE
            #       AND pt_parent.active = TRUE
            # """, (product_ids,))

            self.env.cr.execute("""
                UPDATE sale_order_line sol
                SET product_id = pp_parent.id
                FROM product_product pp
                JOIN product_template pt ON pp.product_tmpl_id = pt.id
                JOIN product_template pt_parent ON pt.parent_template_id = pt_parent.id
                JOIN product_product pp_parent ON pp_parent.product_tmpl_id = pt_parent.id
                WHERE sol.product_id = pp.id
                  AND pp_parent.id IS NOT NULL
                  AND sol.product_id != pp_parent.id
                  AND pt.id IN %s
            """, (product_ids,))


            _logger.info("Updated %s sale.order.line rows", self.env.cr.rowcount)

            # ---------------- PURCHASE ORDER LINE ----------------
            # self.env.cr.execute("""
            #     UPDATE purchase_order_line pol
            #     SET product_id = pp_parent.id
            #     FROM product_product pp
            #     JOIN product_template pt ON pp.product_tmpl_id = pt.id
            #     JOIN product_template pt_parent ON pt.parent_template_id = pt_parent.id
            #     JOIN product_product pp_parent ON pp_parent.product_tmpl_id = pt_parent.id
            #     WHERE pol.product_id = pp.id
            #       AND pp_parent.id IS NOT NULL
            #       AND pol.product_id != pp_parent.id
            #       AND pt.id IN %s
            #       AND pt.uom_category_id IS NULL
            #       AND pt.active = TRUE
            #       AND pt_parent.active = TRUE
            # """, (product_ids,))
            self.env.cr.execute("""
                UPDATE purchase_order_line pol
                SET product_id = pp_parent.id
                FROM product_product pp
                JOIN product_template pt ON pp.product_tmpl_id = pt.id
                JOIN product_template pt_parent ON pt.parent_template_id = pt_parent.id
                JOIN product_product pp_parent ON pp_parent.product_tmpl_id = pt_parent.id
                WHERE pol.product_id = pp.id
                  AND pp_parent.id IS NOT NULL
                  AND pol.product_id != pp_parent.id
                  AND pt.id IN %s
            """, (product_ids,))

            _logger.info("Updated %s purchase.order.line rows", self.env.cr.rowcount)

            # ---------------- STOCK MOVE ----------------
            # self.env.cr.execute("""
            #     UPDATE stock_move sm
            #     SET product_id = pp_parent.id
            #     FROM product_product pp
            #     JOIN product_template pt ON pp.product_tmpl_id = pt.id
            #     JOIN product_template pt_parent ON pt.parent_template_id = pt_parent.id
            #     JOIN product_product pp_parent ON pp_parent.product_tmpl_id = pt_parent.id
            #     WHERE sm.product_id = pp.id
            #       AND pp_parent.id IS NOT NULL
            #       AND sm.product_id != pp_parent.id
            #       AND pt.id IN %s
            #       AND pt.uom_category_id IS NULL
            #       AND pt.active = TRUE
            #       AND pt_parent.active = TRUE
            # """, (product_ids,))
            self.env.cr.execute("""
                UPDATE stock_move sm
                SET product_id = pp_parent.id
                FROM product_product pp
                JOIN product_template pt ON pp.product_tmpl_id = pt.id
                JOIN product_template pt_parent ON pt.parent_template_id = pt_parent.id
                JOIN product_product pp_parent ON pp_parent.product_tmpl_id = pt_parent.id
                WHERE sm.product_id = pp.id
                  AND pp_parent.id IS NOT NULL
                  AND sm.product_id != pp_parent.id
                  AND pt.id IN %s
            """, (product_ids,))

            _logger.info("Updated %s stock.move rows", self.env.cr.rowcount)

            # ---------------- STOCK MOVE LINE ----------------
            # self.env.cr.execute("""
            #     UPDATE stock_move_line sml
            #     SET product_id = pp_parent.id
            #     FROM product_product pp
            #     JOIN product_template pt ON pp.product_tmpl_id = pt.id
            #     JOIN product_template pt_parent ON pt.parent_template_id = pt_parent.id
            #     JOIN product_product pp_parent ON pp_parent.product_tmpl_id = pt_parent.id
            #     WHERE sml.product_id = pp.id
            #       AND pp_parent.id IS NOT NULL
            #       AND sml.product_id != pp_parent.id
            #       AND pt.id IN %s
            #       AND pt.uom_category_id IS NULL
            #       AND pt.active = TRUE
            #       AND pt_parent.active = TRUE
            # """, (product_ids,))
            self.env.cr.execute("""
                UPDATE stock_move_line sml
                SET product_id = pp_parent.id
                FROM product_product pp
                JOIN product_template pt ON pp.product_tmpl_id = pt.id
                JOIN product_template pt_parent ON pt.parent_template_id = pt_parent.id
                JOIN product_product pp_parent ON pp_parent.product_tmpl_id = pt_parent.id
                WHERE sml.product_id = pp.id
                  AND pp_parent.id IS NOT NULL
                  AND sml.product_id != pp_parent.id
                  AND pt.id IN %s
            """, (product_ids,))


            _logger.info("Updated %s stock.move.line rows", self.env.cr.rowcount)

            # ---------------- UPDATE OFFSET ----------------
            total_products = self.env['product.template'].search_count([])
            new_offset = last_offset + batch_size
            if new_offset >= total_products:
                new_offset = 0  # restart from beginning

            Param.set_param('product_update_last_offset', str(new_offset))
            _logger.info("Batch complete. Updated offset to %s", new_offset)

        except Exception as e:
            _logger.error("Error updating product_id history: %s", str(e))
            raise ValidationError(_("Server action failed: %s") % str(e))

    @api.model
    def action_archive_child_products(self):
        """Archive child products and return newly archived + already archived"""

        # Step 1: Fetch products that should be archived now
        self.env.cr.execute("""
            SELECT pt.id, pt.name, pt.default_code, uom.name, categ.name, parent.id, parent.name AS parent_name
            FROM product_template pt
            JOIN uom_uom uom ON uom.id = pt.uom_id
            JOIN uom_category categ ON categ.id = pt.uom_category_id
            JOIN product_template parent ON parent.id = pt.parent_template_id
            WHERE pt.id != pt.parent_template_id
              AND pt.parent_template_id IS NOT NULL
              AND pt.uom_category_id IS NOT NULL
              AND pt.active = TRUE
        """)
        products_to_archive = self.env.cr.fetchall()

        # Step 2: Archive them (only if any)
        if products_to_archive:
            ids_to_archive = tuple([p[0] for p in products_to_archive])
            self.env.cr.execute("""
                UPDATE product_template
                SET active = FALSE
                WHERE id IN %s
            """, (ids_to_archive,))

        # Step 3: Fetch already archived ones
        self.env.cr.execute("""
            SELECT pt.id, pt.name, pt.default_code, uom.name, categ.name, parent.id, parent.name AS parent_name
            FROM product_template pt
            JOIN uom_uom uom ON uom.id = pt.uom_id
            JOIN uom_category categ ON categ.id = pt.uom_category_id
            JOIN product_template parent ON parent.id = pt.parent_template_id
            WHERE pt.id != pt.parent_template_id
              AND pt.parent_template_id IS NOT NULL
              AND pt.uom_category_id IS NOT NULL
              AND pt.active = FALSE
        """)
        already_archived = self.env.cr.fetchall()

        # Step 4: Return newly archived if present, else already archived
        if products_to_archive:
            return products_to_archive
        else:
            return already_archived

    def _generate_archived_products_xlsx(self, archived_products):
        """Generate Excel file for archived products using xlsxwriter (built-in)"""
        fp = BytesIO()
        workbook = xlsxwriter.Workbook(fp, {'in_memory': True})
        worksheet = workbook.add_worksheet("Archived Products")

        # Set default column width for all columns (0 = first col, len(headers)-1 = last col)
        headers = ["ID", "Name", "Internal Reference", "UOM", "UOM Category", "Parent ID" ,"Parent Template"]
        worksheet.set_column(0, len(headers)-1, 20)  # fixed width = 20

        # Headers with bold format
        header_format = workbook.add_format({'bold': True, 'align': 'center'})
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        # Rows
        row = 1
        for prod in archived_products:
            worksheet.write(row, 0, prod[0])              # id
            worksheet.write(row, 1, prod[1])              # name
            worksheet.write(row, 2, prod[2] or "")        # default_code
            worksheet.write(row, 3, prod[3])              # UOM
            worksheet.write(row, 4, prod[4])              # UOM category
            worksheet.write(row, 5, prod[5])              # parent_template_id
            worksheet.write(row, 6, prod[6])              # parent name
            row += 1

        workbook.close()
        fp.seek(0)

        file_data = base64.b64encode(fp.read())

        # Attach without linking to a specific product
        attachment = self.env['ir.attachment'].create({
            'name': 'archived_products.xlsx',
            'type': 'binary',
            'datas': file_data,
            'res_model': 'product.template',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return attachment

    @api.model
    def action_archive_and_export_xlsx(self):
        archived_products = self.action_archive_child_products()
        attachment = self._generate_archived_products_xlsx(archived_products)

        # For button: return file download
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_update_uom_history(self):
        try:
            Param = self.env['ir.config_parameter'].sudo()
            batch_size = 8000

            # Get last offset, default 0
            last_offset = int(Param.get_param('uom_update_last_offset', default=0))

            # Fetch next batch of product templates
            products = self.env['product.template'].search([], offset=last_offset, limit=batch_size)
            if not products:
                last_offset = 0
                products = self.env['product.template'].search([], offset=last_offset, limit=batch_size)

            if not products:
                _logger.info("No product templates found for update.")
                return

            product_template_ids = tuple(products.ids)
            _logger.info("Processing product templates from %s to %s", last_offset + 1, last_offset + len(products))

            # ---------------------
            # Update stock_move_line with JOIN
            # ---------------------
            self.env.cr.execute("""
                UPDATE stock_move_line sml
                SET product_uom_id = pt.uom_id
                FROM product_product pp
                JOIN product_template pt ON pp.product_tmpl_id = pt.id
                WHERE sml.product_id = pp.id
                  AND pp.active = TRUE
                  AND pt.id IN %s
            """, (product_template_ids,))

            # ---------------------
            # Update stock_move with JOIN
            # ---------------------
            self.env.cr.execute("""
                UPDATE stock_move sm
                SET product_uom = pt.uom_id
                FROM product_product pp
                JOIN product_template pt ON pp.product_tmpl_id = pt.id
                WHERE sm.product_id = pp.id
                  AND pp.active = TRUE
                  AND pt.id IN %s
            """, (product_template_ids,))

            # Update the offset
            total_products = self.env['product.template'].search_count([])
            new_offset = last_offset + batch_size
            if new_offset >= total_products:
                new_offset = 0
            Param.set_param('uom_update_last_offset', new_offset)

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'success',
                    'message': _("Server action run successfully"),
                    'next': {'type': 'ir.actions.act_window_close'},
                }
            }

        except Exception as e:
            _logger.error("Error updating UOM history: %s", str(e))
            raise ValidationError(_("Server action failed: %s") % str(e))

    def action_update_po_so_line_uom_history(self):
        try:
            Param = self.env['ir.config_parameter'].sudo()
            batch_size = 10000

            # Get last offset, default 0
            last_offset = int(Param.get_param('uom_update_last_offset') or 0)

            # Fetch next batch of products
            products = self.env['product.template'].search([], offset=last_offset, limit=batch_size)
            if not products:
                last_offset = 0
                products = self.env['product.template'].search([], offset=last_offset, limit=batch_size)

            if not products:
                _logger.info("No product templates found for update.")
                return

            product_ids = tuple(products.ids)
            start_num = last_offset + 1
            end_num = last_offset + len(product_ids)
            _logger.info("Processing product templates from %s to %s", start_num, end_num)

            # Update SALE ORDER LINE
            _logger.info("Updating sale.order.line UOMs")
            self.env.cr.execute("""
                UPDATE sale_order_line sol
                SET product_uom = pt.uom_id
                FROM product_product pp
                JOIN product_template pt ON pp.product_tmpl_id = pt.id
                WHERE sol.product_id = pp.id
                  AND pp.active = TRUE
                  AND pt.id IN %s
            """, (product_ids,))

            # Update PURCHASE ORDER LINE
            _logger.info("Updating purchase.order.line UOMs")
            self.env.cr.execute("""
                UPDATE purchase_order_line pol
                SET product_uom = pt.uom_po_id
                FROM product_product pp
                JOIN product_template pt ON pp.product_tmpl_id = pt.id
                WHERE pol.product_id = pp.id
                  AND pp.active = TRUE
                  AND pt.id IN %s
            """, (product_ids,))

            # Update offset
            total_products = self.env['product.template'].search_count([])
            new_offset = last_offset + batch_size
            if new_offset >= total_products:
                new_offset = 0
            Param.set_param('uom_update_last_offset', str(new_offset))
            _logger.info("Updated offset to %s for next batch", new_offset)

        except Exception as e:
            _logger.error("Error updating PO/SO UOM history: %s", str(e))
            raise ValidationError(_("Server action failed: %s") % str(e))

    def action_update_barcode_and_description(self):
        try:
            Param = self.env['ir.config_parameter'].sudo()
            batch_size = 8000

            # Get last offset, default 0
            last_offset = int(Param.get_param('uom_update_last_offset') or 0)

            # Fetch next batch of products
            products = self.env['product.template'].search([], offset=last_offset, limit=batch_size)
            if not products:
                last_offset = 0
                products = self.env['product.template'].search([], offset=last_offset, limit=batch_size)

            if not products:
                _logger.info("No product templates found for update.")
                return

            product_ids = tuple(products.ids)
            start_num = last_offset + 1
            end_num = last_offset + len(product_ids)
            _logger.info("Processing product templates from %s to %s", start_num, end_num)
            
            for product in products:
                if not product.parent_template_id or not product.parent_template_id.barcode_uom_ids:
                    continue

                for bu in product.parent_template_id.barcode_uom_ids:

                    # Stock Move
                    self.env.cr.execute("""
                        UPDATE stock_move sm
                        SET x_scanned_barcode = %s,
                            description_picking = %s
                        FROM product_product pp
                        WHERE sm.product_id = pp.id
                          AND pp.product_tmpl_id = %s
                          AND sm.product_uom = %s
                    """, (bu.barcode, bu.description, product.id, bu.uom_id.id))

                    # Stock Move Line
                    self.env.cr.execute("""
                        UPDATE stock_move_line sml
                        SET x_scanned_barcode = %s,
                            description = %s
                        FROM product_product pp
                        WHERE sml.product_id = pp.id
                          AND pp.product_tmpl_id = %s
                          AND sml.product_uom_id = %s
                    """, (bu.barcode, bu.description, product.id, bu.uom_id.id))

            # Update offset
            total_products = self.env['product.template'].search_count([])
            new_offset = last_offset + batch_size
            if new_offset >= total_products:
                new_offset = 0
            Param.set_param('uom_update_last_offset', str(new_offset))
            _logger.info("Updated offset to %s for next batch", new_offset)

        except Exception as e:
            _logger.error("Error updating PO/SO UOM history: %s", str(e))
            raise ValidationError(_("Server action failed: %s") % str(e))

    def action_update_barcode_and_description_po_so(self):
        try:
            Param = self.env['ir.config_parameter'].sudo()
            batch_size = 6000

            # Get last offset, default 0
            last_offset = int(Param.get_param('uom_update_last_offset') or 0)
            products = self.env['product.template'].search([('active', '=', True)],offset=last_offset,limit=batch_size)

            # Fetch next batch of products
            # products = self.env['product.template'].search([], offset=last_offset, limit=batch_size)
            if not products:
                last_offset = 0
                products = self.env['product.template'].search([('active', '=', True)],offset=last_offset,limit=batch_size)
                # products = self.env['product.template'].search([], offset=last_offset, limit=batch_size)

            if not products:
                _logger.info("No product templates found for update.")
                return

            product_ids = tuple(products.ids)
            start_num = last_offset + 1
            end_num = last_offset + len(product_ids)
            _logger.info("Processing product templates from %s to %s", start_num, end_num)
            
            for product in products:
                if not product.parent_template_id or not product.parent_template_id.barcode_uom_ids:
                    continue

                for bu in product.parent_template_id.barcode_uom_ids:
                    # Sale order line
                    self.env.cr.execute("""
                        UPDATE sale_order_line sol
                        SET x_scanned_barcode = %s,
                            name = %s
                        FROM product_product pp
                        WHERE sol.product_id = pp.id
                          AND pp.product_tmpl_id = %s
                          AND sol.product_uom = %s
                    """, (bu.barcode, bu.description, product.id, bu.uom_id.id))

                    # Purchase order Line
                    self.env.cr.execute("""
                        UPDATE purchase_order_line pol
                        SET x_scanned_barcode = %s,
                            name = %s
                        FROM product_product pp
                        WHERE pol.product_id = pp.id
                          AND pp.product_tmpl_id = %s
                          AND pol.product_uom = %s
                    """, (bu.barcode, bu.description, product.id, bu.uom_id.id))

            _logger.info("Updated successfully")
            # Update offset
            total_products = self.env['product.template'].search_count([('active', '=', True)])
            new_offset = last_offset + batch_size
            if new_offset >= total_products:
                new_offset = 0
            Param.set_param('uom_update_last_offset', str(new_offset))
            _logger.info("Updated offset to %s for next batch", new_offset)

        except Exception as e:
            _logger.error("Error updating PO/SO UOM history: %s", str(e))
            raise ValidationError(_("Server action failed: %s") % str(e))

    def action_update_manufacturing_history(self):
        try:
            Param = self.env['ir.config_parameter'].sudo()
            ProductTemplate = self.env['product.template']
            ProductBarcodeUom = self.env['product.barcode.uom']

            batch_size = 10000
            last_offset = int(Param.get_param('uom_update_last_offset') or 0)

            # Fetch next batch of products
            products = ProductTemplate.search([], offset=last_offset, limit=batch_size)
            if not products:
                last_offset = 0
                products = ProductTemplate.search([], offset=last_offset, limit=batch_size)

            if not products:
                _logger.info("No product templates found for update.")
                return

            product_ids = tuple(products.ids)
            start_num, end_num = last_offset + 1, last_offset + len(product_ids)
            _logger.info("Processing product templates from %s to %s", start_num, end_num)

            # ---------------------- STOCK MOVE (MRP PRODUCTION) ----------------------
            stock_moves = self.env['stock.move'].search([
                ('raw_material_production_id', '!=', False),
                ('product_id.product_tmpl_id', 'in', product_ids),
            ])
            for move in stock_moves:
                product = move.product_id
                tmpl = product.product_tmpl_id
                parent_tmpl = tmpl.parent_template_id or tmpl  

                barcode_line = self.env['product.barcode.uom'].search([
                    ('product_id', '=', parent_tmpl.id),
                    ('uom_id', '=', move.product_uom.id)
                ], limit=1)

                move.x_scanned_barcode = barcode_line.barcode if barcode_line else False

            # ---------------------- MRP PRODUCTION ----------------------
            self.env.cr.execute("""
                UPDATE mrp_production mp
                   SET product_uom_id = pt.uom_id,
                       product_id = pp.id
                  FROM product_product pp
                  JOIN product_template pt ON pp.product_tmpl_id = pt.id
                 WHERE mp.product_id = pp.id
                   AND pp.active = TRUE
                   AND pt.id IN %s;
            """, (product_ids,))

            productions = self.env['mrp.production'].search([
                ('product_id.product_tmpl_id', 'in', product_ids),
            ])
            for prod in productions:
                tmpl = prod.product_id.product_tmpl_id
                parent_tmpl = tmpl.parent_template_id or tmpl
                barcode_line = ProductBarcodeUom.search([
                    ('product_id', '=', parent_tmpl.id),
                    ('uom_id', '=', prod.product_uom_id.id)
                ], limit=1)
                prod.x_scanned_barcode = barcode_line.barcode if barcode_line else False

            # ---------------------- MRP BOM ----------------------
            self.env.cr.execute("""
                UPDATE mrp_bom mb
                   SET product_uom_id = pt.uom_id,
                       product_tmpl_id = pt.id
                  FROM product_template pt
                 WHERE mb.product_tmpl_id = pt.id
                   AND pt.active = TRUE
                   AND pt.id IN %s;
            """, (tuple(product_ids),))

            boms = self.env['mrp.bom'].search([
                ('product_tmpl_id', 'in', product_ids),
            ])
            for bom in boms:
                tmpl = bom.product_tmpl_id
                parent_tmpl = tmpl.parent_template_id or tmpl
                barcode_line = ProductBarcodeUom.search([
                    ('product_id', '=', parent_tmpl.id),
                    ('uom_id', '=', bom.product_uom_id.id)
                ], limit=1)
                bom.x_scanned_barcode = barcode_line.barcode if barcode_line else False

            # ---------------------- MRP BOM LINE ----------------------
            self.env.cr.execute("""
                UPDATE mrp_bom_line mbl
                   SET product_uom_id = pt.uom_id,
                       product_id = pp.id
                  FROM product_product pp
                  JOIN product_template pt ON pp.product_tmpl_id = pt.id
                 WHERE mbl.product_id = pp.id
                   AND pp.active = TRUE
                   AND pt.id IN %s;
            """, (product_ids,))

            bom_lines = self.env['mrp.bom.line'].search([
                ('product_id.product_tmpl_id', 'in', product_ids),
            ])
            for line in bom_lines:
                tmpl = line.product_id.product_tmpl_id
                parent_tmpl = tmpl.parent_template_id or tmpl
                barcode_line = ProductBarcodeUom.search([
                    ('product_id', '=', parent_tmpl.id),
                    ('uom_id', '=', line.product_uom_id.id)
                ], limit=1)
                line.x_scanned_barcode = barcode_line.barcode if barcode_line else False
                line.name = barcode_line.description if barcode_line else False

            # ---------------------- OFFSET UPDATE ----------------------
            total_products = ProductTemplate.search_count([])
            new_offset = last_offset + batch_size
            if new_offset >= total_products:
                new_offset = 0

            Param.set_param('uom_update_last_offset', str(new_offset))
            _logger.info("Updated offset to %s for next batch", new_offset)

        except Exception as e:
            _logger.error("Error updating PO/SO/MO/MRP/BOM UOM history: %s", str(e))
            raise ValidationError(_("Server action failed: %s") % str(e))

    def action_generate_skipped_barcode_lines(self):
        self.archive_product_fromsheet()
        products = self.env["product.template"].search([])
        BarcodeUom = self.env['product.barcode.uom']
        data = []
        data2 = []
        data3 = []
        data4 = []

        for parent in products:
            # ---------------- Barcode Skipped Logic ---------------- #
            if parent.parent_template_id and parent.id == parent.parent_template_id.id:
                if not parent.barcode_uom_ids:
                    barcode_lines = []

                    # Parent barcode
                    if parent.barcode and not BarcodeUom.search([('barcode', '=', parent.barcode)], limit=1):
                        barcode_lines.append((0, 0, {
                            'product_id': parent.id,
                            'uom_category_id': parent.uom_category_id.id,
                            'description': parent.name,
                            'uom_id': parent.uom_id.id,
                            'barcode': parent.barcode,
                            'sale_price': parent.list_price,
                        }))

                    # Parent multi-barcodes
                    for bc in parent.multi_barcode_ids:
                        barcode_lines.append((0, 0, {
                            'product_id': parent.id,
                            'uom_category_id': parent.uom_category_id.id,
                            'description': parent.name,
                            'uom_id': parent.uom_id.id,
                            'barcode': bc.multi_barcode,
                            'sale_price': parent.list_price,
                        }))

                    # Child products
                    child_products = self.env['product.template'].search([
                        ('parent_template_id', '=', parent.id),
                        ('id', '!=', parent.id)
                    ])
                    for child in child_products:
                        is_child = child.id != child.parent_template_id.id
                        if is_child and child.parent_template_id and child.parent_template_id.active:
                        # if is_child and child.parent_template_id and child.parent_template_id.active and child.product_variant_id.active:
                            parent = child.parent_template_id
                            comp_id = child.company_id.id or self.env.user.company_id.id
                            chatter_lines = []

                            # Get child product quants
                            self.env.cr.execute("""
                                SELECT location_id, quantity
                                FROM stock_quant
                                WHERE product_id = %s
                            """, (child.product_variant_id.id,))
                            child_quants = self.env.cr.fetchall()

                            for loc_id, qty in child_quants:
                                adjusted_qty = child.unit_qty * qty if child.unit_qty else qty

                                # Parent old qty
                                self.env.cr.execute("""
                                    SELECT quantity
                                    FROM stock_quant
                                    WHERE product_id = %s
                                      AND location_id = %s
                                      AND company_id = %s
                                """, (parent.product_variant_id.id, loc_id, comp_id))
                                res = self.env.cr.fetchone()
                                parent_old_qty = res[0] if res else 0.0

                                # Update or insert
                                self.env.cr.execute("""
                                    UPDATE stock_quant
                                    SET quantity = quantity + %s
                                    WHERE product_id = %s
                                      AND location_id = %s
                                      AND company_id = %s
                                """, (adjusted_qty, parent.product_variant_id.id, loc_id, comp_id))

                                if self.env.cr.rowcount == 0:
                                    self.env.cr.execute("""
                                        INSERT INTO stock_quant (product_id, location_id, quantity, reserved_quantity, company_id)
                                        VALUES (%s, %s, %s, 0, %s)
                                    """, (parent.product_variant_id.id, loc_id, adjusted_qty, comp_id))
                                    parent_old_qty = 0.0

                                # Sync supplier UoM
                                for seller in child.seller_ids:
                                    if child.uom_po_id.id != seller.product_uom.id:
                                        seller.write({'product_uom': child.uom_po_id.id})
                                        # seller.product_uom = child.uom_po_id

                                # Chatter logging
                                location_rec = self.env['stock.location'].browse(loc_id)
                                if location_rec.usage == 'internal':
                                    chatter_lines.append(
                                        f"Location: <b>{location_rec.display_name}</b><br/>"
                                        f"Parent Qty (Before): {parent_old_qty}<br/>"
                                        f"Child Qty × Unit Qty: {adjusted_qty} "
                                        f"({qty} × {child.unit_qty if child.unit_qty else 1})<br/>"
                                        f"<b>Total Qty (After):</b> {parent_old_qty + adjusted_qty}"
                                    )

                            # Post chatter
                            if chatter_lines:
                                parent.message_post(
                                    body="<br/><br/>".join(chatter_lines),
                                    subtype_xmlid="mail.mt_note"
                                )

                            # Ensure child seller UoM is correct
                            for seller in child.seller_ids:
                                if child.uom_po_id.id != seller.product_uom.id:
                                    seller.write({'product_uom': child.uom_po_id.id})
                                    # seller.product_uom = child.uom_po_id
                                    
                        if child.barcode:
                            barcode_lines.append((0, 0, {
                                'product_id': parent.id,
                                'uom_category_id': child.uom_category_id.id,
                                'description': f"{parent.name} X {child.uom_id.name}",
                                'uom_id': child.uom_id.id,
                                'barcode': child.barcode,
                                'sale_price': child.list_price,
                            }))
                        for bc in child.multi_barcode_ids:
                            barcode_lines.append((0, 0, {
                                'product_id': parent.id,
                                'uom_category_id': child.uom_category_id.id,
                                'description': f"{parent.name} X {child.uom_id.name}",
                                'uom_id': child.uom_id.id,
                                'barcode': bc.multi_barcode,
                                'sale_price': child.list_price,
                            }))

                    # Update parent record in DB
                    if barcode_lines:
                        parent.write({'barcode_uom_ids': barcode_lines})

                # Sheet 4 → Parent without barcode lines
                if not parent.barcode_uom_ids and parent.type == 'product':
                    data4.append([
                        parent.id,
                        parent.default_code or "",
                        parent.barcode or "",
                        parent.name,
                        parent.parent_template_id.name,
                        parent.unit_qty,
                        parent.ctn_qty,
                        parent.uom_id.category_id.name if parent.uom_id else "",
                        parent.parent_template_id.uom_id.category_id.name if parent.parent_template_id.uom_id else "",
                        parent.uom_id.name if parent.uom_id else "",
                        parent.parent_template_id.uom_id.name if parent.parent_template_id.uom_id else "",
                    ])

            # ---------------- Sheet 1 Logic ---------------- #
            if parent.uom_id and parent.uom_id.category_id:
                try:
                    category_val = int(''.join([c for c in parent.uom_id.category_id.name if c.isdigit()]))
                    uom_val = int(''.join([c for c in parent.uom_id.name if c.isdigit()]))

                    if uom_val > category_val:
                        data.append([
                            parent.id,
                            parent.default_code or "",
                            parent.barcode or "",
                            parent.name,
                            parent.parent_template_id.name if parent.parent_template_id else "",
                            parent.unit_qty,
                            parent.ctn_qty,
                            parent.uom_id.category_id.name,
                            parent.parent_template_id.uom_id.category_id.name if parent.parent_template_id else "",
                            parent.uom_id.name,
                            parent.parent_template_id.uom_id.name if parent.parent_template_id else "",
                        ])
                except Exception:
                    pass

            # ---------------- Sheet 2 Logic ---------------- #
            # if parent.unit_qty == 1 and parent.id != (parent.parent_template_id.id if parent.parent_template_id else 0):
            #     data2.append([
            #         parent.id,
            #         parent.default_code or "",
            #         parent.barcode or "",
            #         parent.name,
            #         parent.parent_template_id.name if parent.parent_template_id else "",
            #         parent.unit_qty,
            #         parent.ctn_qty,
            #         parent.uom_id.category_id.name if parent.uom_id else "",
            #         parent.parent_template_id.uom_id.category_id.name if parent.parent_template_id else "",
            #         parent.uom_id.name if parent.uom_id else "",
            #         parent.parent_template_id.uom_id.name if parent.parent_template_id else "",
            #     ])

            # ---------------- Sheet 3 Logic ---------------- #
            if parent.parent_template_id and not parent.parent_template_id.active and parent.type == 'product':
                data3.append([
                    parent.id,
                    parent.default_code or "",
                    parent.barcode or "",
                    parent.name,
                    parent.parent_template_id.name,
                    parent.unit_qty,
                    parent.ctn_qty,
                    parent.uom_id.category_id.name if parent.uom_id else "",
                    parent.parent_template_id.uom_id.category_id.name if parent.parent_template_id.uom_id else "",
                    parent.uom_id.name if parent.uom_id else "",
                    parent.parent_template_id.uom_id.name if parent.parent_template_id.uom_id else "",
                ])

        # ---------------- Excel File Creation ---------------- #
        wb = Workbook()
        ws = wb.active
        ws.title = "UoM Greater Report"

        headers = ["ID", "Internal Reference", "Barcode", "Product Name", "Parent Product", "Unit Quantity", "CTN Quantity",
                   "UoM Category", "Parent UoM Category", "UoM", "Parent UoM"]
        ws.append(headers)
        for row in data:
            ws.append(row)
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 25

        # Sheet 2
        # ws2 = wb.create_sheet(title="Product with 1X uom")
        # ws2.append(headers)
        # for row in data2:
        #     ws2.append(row)
        # for i in range(1, len(headers) + 1):
        #     ws2.column_dimensions[get_column_letter(i)].width = 25

        # Sheet 3
        ws3 = wb.create_sheet(title="Archived Parent Products")
        ws3.append(headers)
        for row in data3:
            ws3.append(row)
        for i in range(1, len(headers) + 1):
            ws3.column_dimensions[get_column_letter(i)].width = 25

        # Sheet 4
        ws4 = wb.create_sheet(title="Parent Multi barcode is not set")
        ws4.append(headers)
        for row in data4:
            ws4.append(row)
        for i in range(1, len(headers) + 1):
            ws4.column_dimensions[get_column_letter(i)].width = 25

        # Save and return
        fp = BytesIO()
        wb.save(fp)
        file_data = base64.b64encode(fp.getvalue())
        fp.close()

        attachment = self.env['ir.attachment'].create({
            'name': "generate_skipped_barcode_lines.xlsx",
            'type': 'binary',
            'datas': file_data,
            'res_model': 'product.template',
            'res_id': 0,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

    @api.model
    def archive_product_fromsheet(self):
        # remove (Picking Operations with barccode) report from print
        report = self.env['ir.actions.report'].search([('name', '=', 'Picking Operations with barccode')])
        if report:
            report.unlink_action()

        module_path = os.path.dirname(os.path.abspath(__file__))
        folder_path = os.path.join(module_path, "..", "UPDATE PARENT TEMPLATE")
        folder_path = os.path.normpath(folder_path)

        if not os.path.exists(folder_path):
            raise UserError("Folder not found: %s" % folder_path)

        for filename in os.listdir(folder_path):
            if filename.endswith((".xlsx", ".xlsm")):
                file_path = os.path.join(folder_path, filename)
                _logger.info("Processing file: %s", file_path)

                try:
                    wb = load_workbook(file_path, read_only=True)
                    ws = wb.active

                    # Assuming header row is 1
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        product_name = row[1]
                        if not product_name:
                            continue

                        product = self.search([("name", "=", product_name)], limit=1)
                        if product:
                            product.write({"active": False})
                            _logger.info("Archived product: %s", product_name)
                        else:
                            _logger.warning("Product not found: %s", product_name)

                except Exception as e:
                    _logger.error("Error reading file %s: %s", file_path, e)


class ProductProduct(models.Model):
    _inherit = 'product.product'

    @api.model
    def create(self, vals):
        product = super(ProductProduct, self).create(vals)
        if not product.active:
            product.active = True

        return product